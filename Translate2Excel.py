
import os
import sys
import xlrd
import openpyxl
from utils import translateYoudao
from utils import translateBaidu
import requests
import sys
from Config import EXCEL_DIR
from Config import EXCEL_EXT
from Config import UNITY_TABLE_FIELD_FILTER
from Config import UnityDataDir

try:
    import vlc
except ImportError:
    pass
except OSError:
    pass  # `pip install python-vlc` needed

try:
    import pyttsx3
except ImportError:
    pass

class Translate2Excel:
    # 构造函数
    def __init__(self):
        self.mExcelFiles = []  # 所有的excel文件
        self.translateDic = {} 

    # 外部处理函数
    def Process(self):
        self.RecursiveSearchExcel(EXCEL_DIR)
        self.ProcessTranslate2Excel()

    # 递归查找文件
    def RecursiveSearchExcel(self, path):
        for pathdir in os.listdir(path):  # 遍历当前目录
            fullpath = os.path.join(path, pathdir)

            if os.path.isdir(fullpath):
                self.RecursiveSearchExcel(fullpath)
            elif os.path.isfile(fullpath):
                if os.path.splitext(fullpath)[1] == EXCEL_EXT:
                    self.mExcelFiles.append(fullpath)

    # 处理excel文件
    def ProcessTranslate2Excel(self):
        allbytesdata = bytes()

        # 处理每个文件
        for filename in self.mExcelFiles:
            data = xlrd.open_workbook(filename)
            table = data.sheets()[0]
            fields = self.FilterFieldData(table, UNITY_TABLE_FIELD_FILTER)

            #增加多语言配置
            languageKeys = ["EN"]
            languageTables = []
            for x in data.sheets():
                if x != table and x.name in languageKeys:
                    languageTables.append(x)
            
            if(len(languageTables)!=0):
                print("\r"+'-'*70)
                print("\r找到多语言表格："+filename)
                wb = openpyxl.load_workbook(filename)
                #目前接口仅支持 英文、日语、韩语、法语
                sheet = wb["EN"]          
                
                for i in range(table.ncols):
                    cell1 = table.cell(3, i).value
                    for j in range(sheet.max_column):
                        cell2 = sheet.cell( 4, j+1).value   
                        if(cell1==cell2):
                            #找到需要翻译的字段 
                            print("\r正在翻译字段："+cell1)
                            for m in range(5,table.nrows):
                                val = table.cell(m, i).value
                                if(val != ""):
                                    sheet.cell( m+1, j+1, self.GetTranslateStr(val)) 
                                    print( "\r进度："+str(m-5)+"/"+str(table.nrows-6)," {:.0f}% ".format( (m-5) /(table.nrows-6) * 100 ), end="")
                                    sys.stdout.flush()
                print("\r翻译完成："+filename)            
                wb.save(filename)


    # 筛选字段数据
    def FilterFieldData(self, table, fieldfilter):
        fields = []
        for index in range(table.ncols):
            row = table.cell(1, index).value
            for field in fieldfilter:
                if row == field:
                    fields.append(index)

        return fields
        
    def GetTranslateStr(self,string):
        s = self.translateDic.get(string,"")
        if(s == ""):
            s = translateBaidu(string)
        self.translateDic[string] = s
        return s
  